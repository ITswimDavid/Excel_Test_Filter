from openpyxl import load_workbook
from PIL import Image
from io import BytesIO
import matplotlib.pyplot as plt
from openpyxl.cell import Cell

from openpyxl_image_loader import SheetImageLoader

# Load the workbook and select the worksheet
wb = load_workbook('your_file.xlsx')
ws = wb.active

# calling the image_loader
image_loader = SheetImageLoader(ws)


# get the image (put the cell you need instead of 'A1')


# showing the image
# image.show()
def is_image_in_cell(cell):
    for image in ws._images:
        anchor = image.anchor
        if anchor.from_col <= cell.column <= anchor.to_col and anchor.from_row <= cell.row <= anchor.to_row:
            return True

    return False


for i in range(1, ws.max_row + 1):
    row = [cell for cell in ws[i]]  # sheet[n] gives nth row (list of cells)

    for cel in row:
        cel: Cell  # use type hinting to get code completion

        cel_coordinate = cel.coordinate
        if image_loader.image_in(cel_coordinate):  # needs coords of cell
            print("Cell: " + cel_coordinate)
            image_loader.image_in(cel_coordinate)
            image = image_loader.get(cel_coordinate)
            image.show()

        print(cel.value)

    print(row)  # list of cell values of this row

# Loop through the worksheets images

def is_png(image_path):
    try:
        img = Image.open(image_path)
        return img.format == 'PNG'
    except IOError:
        return False

image_path = 'Image.png'
if is_png(image_path):
    print("The image is in PNG format.")
else:
    print("The image is not in PNG format.")


def convert_to_png(image_path, output_path):
    try:
        img = Image.open(image_path)
        if img.format != 'PNG':
            img.save(output_path, 'PNG')
            print("Image converted to PNG format.")
        else:
            print("The image is already in PNG format.")
    except IOError:
        print("Unable to open the image.")

image_path = 'Image.jpg'
output_path = 'output_image.png'
convert_to_png(image_path, output_path)